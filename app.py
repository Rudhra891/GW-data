import streamlit as st
import pandas as pd
from io import BytesIO
from datetime import date

import os
from openpyxl import load_workbook

import pandas as pd
from io import BytesIO
from openpyxl import load_workbook

image_url1 = "https://www.egr.msu.edu/hydrology/Home/hydro/Hydro_cycle_02.gif"
image_url2 = "https://media1.giphy.com/media/v1.Y2lkPTc5MGI3NjExdHN5MHNxdzNmYmlvZTd1NDdraGdqamJtODNpcWxlc2gyMHV6d3plcCZlcD12MV9pbnRlcm5hbF9naWZfYnlfaWQmY3Q9Zw/Bdlzpicqzu5SmLswrH/giphy.gif"
image_url3 ="https://i.makeagif.com/media/11-10-2019/8XfJT6.gif"

html = f"""
<div style="display: flex; flex-direction: column; align-items: center;">
  <img src="{image_url1}" alt="Hydrological Cycle" style="width: 120%; margin-bottom: 40px;" />
  <img src="{image_url2}" alt="Geoph Data" style="width: 120%; margin-bottom: 40px;" />
  <img src="{image_url3}" alt="Geoph Data" style="width: 120%;" />
  
</div>
"""

st.sidebar.markdown(html, unsafe_allow_html=True)



st.markdown("""
<style>
body, .block-container {
    margin: 0;
    padding: 0;
}
.header {
    position: fixed; top: 0; left: 0; width: 100%;
    height: 100px;
    background: linear-gradient(90deg,#FFE53B, #FFCE00);
    color: white;
    display: flex; align-items: center; justify-content: space-between;
    padding: 10px 40px;
    box-sizing: border-box; z-index: 1000;
    font-size: 1.5rem; font-weight: bold;
}
.footer {
    position: fixed; bottom: 0.5; left: 3; width: 50%;
    height: 30px; /* Reduced height */
    background: linear-gradient(90deg, .#FFE53B, #FFCE00);
    color: white; display: flex; align-items: center; justify-content: center;
    padding: 10px 20px; box-sizing: b..order-box; z-index: 1000;
    font-size: 0.5rem;
}
.content {
    padding-top: 120px; /* Increased top padding */
    padding-bottom: 80px; /* Increased bottom padding */
}
</style>

<div class="header">
---
    <nav>
        <a href="#" style="color: white; text-decoration: none; margin: 0 15px;">Home</a>
        <a href="#" style="color: white; text-decoration: none; margin: 0 15px;">About</a>
        <a href="#" style="color: white; text-decoration: none; margin: 0 15px;">Contact</a>
    </nav>
</div>

<div class="footer">© 2019 Bebpl | All Rights Reserved</div>

<div class="content">
""", unsafe_allow_html=True)


st.markdown("</div>", unsafe_allow_html=True)

st.markdown("""
    <style>
        /* The bar container */
        .custom-bar {
            background-color: #D3F527;        /* Bar color: change this */
            position: absolute;                /* Fixed position on viewport */
            top: 4.7cm;                        /* Vertical position from the top */
            left: -10%;                      /* Horizontal position from left */
            width: 120%;                     /* Width of the bar */
            height: 0.4cm;                  /* Height (thickness) of the bar */
            z-index: 999;                   /* Keeps the bar on top */
        }

        /* Optional: add a subtle shadow for depth */
        .custom-bar-shadow {
            box-shadow: 0px 0px 8px rgba(0, 0, 0, 0.2);
        }
    </style>

    <!-- The actual bar element -->
    <div class="custom-bar custom-bar-shadow"></div>
    """, unsafe_allow_html=True)





def save_to_excel(df: pd.DataFrame, file_path="data.xlsx") -> bytes:
    try:
        book = load_workbook(file_path)
    except FileNotFoundError:
        book = None

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        if book:
            writer._book = book
            if "Sheet1" in book.sheetnames:
                del book["Sheet1"]
        df.to_excel(writer, index=False, sheet_name="Sheet1")

    return output.getvalue()



def main():
     
    LOGO_URL = "https://bebpl.com/wp-content/uploads/2023/07/BLUE-ENERGY-lFINAL-LOGO.png"
    
    st.markdown(
    f'<div style="text-align:center;"><img src="{LOGO_URL}" style="max-width:350px; height:10; padding-top: 0.005cm;",></div>',
    unsafe_allow_html=True
)
    st.text("")
    st.text("")
    st.title("DATA FORM-GROUNDWATER SURVEY-REPORT")  
    st.text("")
    st.markdown(
    """
    <style>
    .stApp {
        background: linear-gradient(to right, #000080, #F5A623, #D0D0D0 ); 
        background-size: 400% 400%;
        animation: candy 25s ease infinite;
        color: black;
    }

    @keyframes candy {
        0% {background-position: 0% 50%;}
        50% {background-position: 100% 50%;}
        100% {background-position: 0% 50%;}
    }
    </style>
    """,
    unsafe_allow_html=True
)







    
    
    
    
    #options
    terrain_options = ["Granitic","Granitoid gneiss","gneiss","charnockite", "khondalite","Basaltic","Limestone","Laterite","Quartzite",
                       "Migmatite","Shale","Schist","Dolerite","Anorthosite / gabbro / dunite","Porphyritic granite","Metabasite","Migmatitic","Alluvial-covered basement"
                      ]
    
    geomorphic_options = ["Pediplain shallow","Pediplain moderate","Pediment","Inselberg complex","Dyke ridge","Plateau slightly dissected",
                       "Plateau moderately dissected","Pediment Inselberg complex","Valley fill shallow","Denudational hill",
                       "Buried Pediplain shallow","Inselberg","Plateau undissected","Pediplain moderate-Pediplain shallow","Pediplain shallow-Pediment",
                       "Pediplain moderate-Pediment","Pediplain shallow-Dyke ridge","Pediment-Inselberg complex","Inselberg-Dyke ridge",
                       "Buried pediplain shallow-Pediplain moderate","Water body-Pediplain moderate"]
    
    soil_type_options =["Black Cotton Soil","Red Soil","Brown soil(dark)","Brown Soil (light)","Laterite soil","Clay"]
    
    des_soil_options = ["Black Cotton Soil typically expansive, fine-grained and shrink-swell prone—low permeability limits infiltration, often causing surface water retention.",
                 "Red Soil generally well-drained and moderately porous—supports decent infiltration but low water-holding capacity due to low clay and organic matter.",
                 "Brown Soils  moderately fertile and better structured—supports moderate infiltration and retains moisture reasonably well.",
                 "Brown Soils  lighter texture with lower organic content—offers fair infiltration but limited water retention.",
                 "Laterite Soils tropical, highly porous and leached—can be coarse-grained, enabling fast infiltration though often degraded.",
                 "Clay, fine-textured, compact and poorly drained—very low permeability restricts groundwater recharge but retains moisture near surface."
                ]
    
    year_rainfall_options = ["2023-2024","2024-2025","2025-2026","2026-2027","2027-2028","2028-2029","2029-2030","2030-2031","2031-2032",
             "2032-2033","2033-2034","2034-2035","2035-2036","2036-2037","2037-2038","2039-2040"]
    water_quality_options = ["good to moderate","moderate to poor","poor","Good"]
    
    grad_trav_direction_options = ["EW","NS","NE-SW","NW-SE"]
    
    type_of_land_options = ["Agricultural / cropland land","Form land","Plot","Barren land","Hilly terrain","Mountainous terrain","Valleys / colluvial fans","Terraces","Gentle to rolling terrain","Forest / scrub"
                           ]
    grad_station_int_options = ["5m","10m","15m","20m"]
    c1c2_options =["200m","300m","400m","500m","600m","700m","800m","900m","1000m"]
    admt_electrode_int_options = ["0.5m","1m","1.5m","2m","2.5m","3m","3.5m","4m","4.5m","5m"]
    yields_options = ["1 - 1.5","1.5 - 2","2 - 2.5","> 1.5","< 1.5"]
    
    st.markdown(
    """
    <style>
    /* Style for input label text (widget labels) */
    .stTextInput > label, 
    .stNumberInput > label,
    .stSlider > label,
    .stSelectbox > label,
    .stRadio > label,
    .stCheckbox > label {
        font-size: 40px !important;
        font-weight: bold !important;
        color: black !important;
    }

    /* Style the input box user text (the typed content) */
    input, textarea {
        font-size: 20px !important;
    }
    </style>
    """,
    unsafe_allow_html=True,
)
    
     
    st.markdown('<div style="background-color:#1f4e79;color:white;padding:5px;border-radius:5px">1. General Information</div>', unsafe_allow_html=True)
    st.text("")
    client_name = st.text_input("CLIENT NAME",placeholder="Shri. Rudra Venkatesh Garu/ Blue Infra Pvt Ltd",help= "please keep Shri. in front of name of the client and end of the name Garu if company directly give name with at the end Pvt Ltd")
    date_selected = st.date_input("SURVEY DATE", value=date.today())
    date_selected = date_selected.strftime("%d-%m-%Y")
    ref_no = st.text_input("Project Name/Reference",help="SN/12/2025",placeholder="SN/12/2025")
    village = st.text_input("VILLAGE")
    mandal = st.text_input("MANDAL")
    district = st.text_input("DISTRICT")
    state = st.text_input("STATE")
    pin = st.text_input("PINCODE")
    area = st.text_input("AREA",help="survey area in terms of Acres/Sq Yards/Sq Feets",value="15 Acres")
    nearest_town = st.text_input("Nearest Town",help="Enter Nearest Town to the survey area")
    to_nearest_town = st.text_input("Distance To Nearest Town",help="in terms of km",value="10 km" )
    nearest_city = st.text_input("Nearest City",help="Enter Nearest City to the survey area")
    to_nearest_city = st.text_input("Distance To Nearest City",help="in terms of km",value="80 km" )
    
    type_of_land = st.selectbox("Type of Land",options=type_of_land_options + ["Other"],help="select from dropdown list")
    if type_of_land == "Other":
        type_of_land_manual = st.text_input("Enter type of land: ")
        if type_of_land_manual:
            type_of_land = type_of_land_manual
    
    st.markdown('<div style="background-color:#1f4e79;color:white;padding:5px;border-radius:5px">2. GEOGRAPHICAL</div>', unsafe_allow_html=True)
    st.text("")
    min_lat = st.text_input("Minimum Latitude",placeholder="17.24574",help="Enter Minimum Latitude of the survey area")
    max_lat = st.text_input("Maximum Latitude",placeholder="17.35574",help="Enter Maximum Latitude of the survey area")
    min_lon = st.text_input("Minimum Longitude",placeholder="79.24574",help="Enter Minimum Longitude of the survey area")
    max_lon = st.text_input("Maximum Longitude",placeholder="79.35574",help="Enter Maximum Longitude of the survey area")
    highest_elevation = st.text_input("Highest Elevation",help="330m in the NW corner",placeholder="330m in the NW corner")
    lowest_elevation = st.text_input("Lowest Elevation",help= "250m in NS Corner",placeholder="250m in NS Corner")
    
    
    rain_fall = st.text_input("Rain fall",value="600 mm")
    
    
    year_rainfall = st.selectbox("Select Year of Rainfall",options=year_rainfall_options +["Other"],help="Select Year of Rain-Fall")    
    if year_rainfall == "Other":
        year_rainfall_manual = st.text_input("Type year_range")
        if year_rainfall_manual:
            year_rainfall = year_rainfall_manual
            
    temp = st.text_input("Temperature",value="15 deg. C – 45 deg. C",help="Temperature in the survey area")
    
    yield_from_prospects = st.text_input("Yield From Prospects", value="30 to 80 m and >80m deep well and 30 to 50 and 100 to 200 LPM yield")
    
    water_quality = st.selectbox("Water Quality",options= water_quality_options + ["Other"],help="select the water quality from the dropdown/type by selection other")
    if water_quality == "Other":
        water_quality_manual = st.text_input("Type water quality ")
        if water_quality_manual:
            water_quality = water_quality_manual
    
    st.markdown('<div style="background-color:#1f4e79;color:white;padding:5px;border-radius:5px">3. Hydro-Geological</div>', unsafe_allow_html=True)
    st.text("")
    h_geo = st.text_area("Describe hydrogeology and existing bores", height=250,
                          value="Hydrogeology investigates the occurrence, movement, and quality of groundwater within the Earth’s subsurface, emphasizing the role of geologic formations in controlling flow and storage. It integrates geological, hydrological, and geochemical principles to understand aquifer characteristics, recharge and discharge processes, and groundwater–surface water interactions. This knowledge is fundamental for assessing water availability, managing resources sustainably, and mitigating contamination risks across diverse geologic settings.",
                          help="this description directly will be inserted as paragrah in the report")
    
    st.markdown('<div style="background-color:#1f4e79;color:white;padding:5px;border-radius:5px">4. GEOPHYSICAL</div>', unsafe_allow_html=True)
    st.text("")
    no_grad_trav = st.text_input("Number of Grad Traverses",help = "Howmany Gradient survey lines?",placeholder=4)
    grad_trav_direction = st.selectbox("Gradient Line Direction",options=grad_trav_direction_options + ["Other"],help="select the gradient survey lines direction from the dropdown list")
    
    if grad_trav_direction == 'Other':
        grad_trav_direction_manual = st.text_input("Type Gradient Line Direction")
        if grad_trav_direction_manual:
            grad_trav_direction = grad_trav_direction_manual
            
    grad_trav_name = st.text_input("Name of Grad Traverses",help="Trav-0, N50, N90 and N100 and S50, S100",placeholder="L0, N50, N100 and S50 and S100")
    
    grad_station_int = st.selectbox("P1P2",options=grad_station_int_options +  ["Other"],help="P1P2 distance in gradient survey, in meters") 
    if grad_station_int == "Other":
        grad_station_int_manual = st.text_input("Eter gradient station interval")
        if grad_station_int_manual:
            grad_station_int = grad_station_int_manual
    
    c1c2 = st.selectbox("C1C2",options=c1c2_options+["Other"],help= "C1C2 distance in gradient survey, in meters")
    if c1c2 == "Other":
        c1c2_manual =st.text_input("Enter c1c2 spacing")
        if c1c2_manual:
            c1c2 = c1c2_manual
    res_highs = st.text_input("Risistivity High Range",help="1200-800",value="1200-800")
    res_high_extends = st.text_input("High Resisvity Extends",value="North to South",help="write in which direction high resistivity trend is moving")
    res_low = st.text_input("Risistivity Low Range",value="300-150")
    res_low_extends = st.text_input("Low Resisvity Extends",value="South west and Nort East parts",help="write in which direction low resistivity trend is moving")
    #res_low_trend_des = st.text_input("Low Resistivity Trends Description",help="in the southwest and northwest part as well in northeast region may reflects  presence of  fracture system in the study area")
    res_low_trend_des = st.text_area("Low Resistivity Trends Description", height=100,
                                      value = "in the southwest and northwest part as well in northeast region may reflects  presence of  fracture system in the study area",help="how this this low resistivity in tems of directions, reflects the fracture systems for groundwater prospects" )
    res_relief = st.text_input("Resistivity- Relief value",placeholder=254)
    
    admt_electrode_int = st.selectbox("ADMT_Electode Spacing",options=admt_electrode_int_options+["Other"],help="select the distance between each electrode")
    if admt_electrode_int == "Other":
        admt_electrode_int_manual = st.text_input("Type AMDT Electode spacing")
        if admt_electrode_int_manual:
            admt_electrode_int = admt_electrode_int_manual
    st.markdown('<div style="background-color:#1f4e79;color:white;padding:5px;border-radius:5px">5. RECCOMENDATIONS</div>', unsafe_allow_html=True)
    st.text("")
    recom_bores = st.text_input("Number of Ground Marking Bore points",placeholder="6",help="howmany points marked in the field/site")
    final_recom_points = st.text_input("Number Of Finalized Bore points",placeholder="3",help="howmany points finalized after analysis in the office")
    recom_points_order = st.text_input("Recommended Points as order as per Priority",help="For example: 2,5,3",placeholder="2,5,3")
    
    yield_ = st.selectbox("Expected Yield",options=yields_options + ["Other"],help="select the expected yield from the dropdown list")
    if yield_ == "Other":
        yield_manual = st.text_input("Enter Expected Yield")
        if yield_manual:
            yield_ = yield_manual
    
    water_zone_depths = st.text_input("Water_Zone_Depths",help="40ft, 120ft, 180ft-240ft, 340ft, 440ft to 480ft",placeholder="40ft, 120ft, 180ft-240ft, 340ft, 440ft to 480ft") 
    considerable_depths = st.text_input("Considerable Depths",help="200 - 1000 feets",placeholder="200 - 1000 feets")
    
        
    st.markdown('<div style="background-color:#1f4e79;color:white;padding:5px;border-radius:5px">6. GEOLOGY & MORPHOLOGY</div>', unsafe_allow_html=True)
    st.text("")
    terrain_type = st.selectbox("Terrain Type", options=terrain_options + ["Other"],help="surface geology of survey area; for example granite/basalt...")
    if terrain_type == "Other":
        terrain_manual = st.text_input("Enter terrain type :")
        if terrain_manual:
            terrain_type = terrain_manual
   
    geomorphic_unit = st.selectbox("Geomorphology",options=geomorphic_options + ["Other"],help="Geomorphology of the survey area")    
    if geomorphic_unit == "Other":
        geomorphic_manual = st.text_input("Enter Geophology: ")
        if geomorphic_manual:
            geomorphic_unit = geomorphic_manual
   
    weathering_depth = st.text_input("weathering_depth",value="shallow")
    comparison_depth = st.text_input("comparison_depth",value="deeper")
    overlying_materials = st.text_input("overlying_materials",value="soils or weathered debris",help= "during survey, what kind of material is there on the surface")
    
    soil_type = st.selectbox("Soil_type",options=soil_type_options + ["Other"],help="What type of soils observed in the survey area")
    if soil_type == "Other":
        soil_type_manual = st.text_input("Enter type of soil")
        if soil_type_manual:
            soil_type = soil_type_manual
            
            
    des_soil = st.selectbox("description of soil", options = des_soil_options + ["Other"],help="select the identified soils description/descibe in the similar way given...")
    if des_soil == "Other":
        des_soil_manual = st.text_input("Type soil description")
        if des_soil_manual:
            des_soil = des_soil_manual
            
    
    figure_number = st.text_input("Figure_number",value="1", help="No need to change this keep it as it is...")
    
    

    #highest_elevation = st.text_input("Highest Elevation (in meters):")
    #highest_elevation_num = None
    #if highest_elevation:
    #    try:
    #        highest_elevation_num = float(highest_elevation)
    #    except ValueError:
    #        st.error("Please enter a valid number for highest elevation.")
            
            
            
    min_lat_num = None
    if min_lat:
        try:
            min_lat_num = float(min_lat)
        except ValueError:
            st.error("Enter a valid number for min_lat.")
            
            
    max_lat_num = None
    if max_lat:
        try:
            max_lat_num = float(max_lat)
        except ValueError:
            st.error("Enter a valid number for min_lat.")
            
    min_lon_num = None
    if min_lon:
        try:
            min_lon_num = float(min_lon)
        except ValueError:
            st.error("Enter a valid number for min_lon.")
            
    max_lon_num = None
    if max_lon:
        try:
            max_lon_num = float(max_lon)
        except ValueError:
            st.error("Enter a valid number for max_lon.")
    

    st.write("---")
    st.write("Image Path : default")
    st.markdown('<div style="background-color:#1f4e79;color:white;padding:5px;border-radius:5px">7. IMAGE PATHS</div>', unsafe_allow_html=True)
    st.text("")
    # Default fixed path
    geology_img_default = r"C:\images\0.jpg"

    # Text input for image path, prefilled with default
    geology_img_path = st.text_input(
        "Geology Image Path:",value=geology_img_default,help="You can edit this if image file path is different",placeholder=geology_img_default
        )
    
    
    home_page_default = r"C:\images\1.jpg"
    home_page_path = st.text_input( "home_page_IMG",value=home_page_default,placeholder=home_page_default)
    
    study_area_default = r"C:\images\2.jpg"
    study_area_path = st.text_input("Study_Area_IMG",value=study_area_default,placeholder=study_area_default)
    
    gw_prospects_default = r"C:\images\3.jpg"
    gw_prospects_path = st.text_input("Ground water Prospects IMG",value=gw_prospects_default,placeholder=gw_prospects_default)
    
    drainage_pattern_default = r"C:\images\4.jpg"
    drainage_pattern_path = st.text_input("Drainage_Pattern_IMG",value=drainage_pattern_default,placeholder=drainage_pattern_default)
    
    lineaments_default =r"C:\images\5.jpg"
    lineaments_path = st.text_input("Lineament IMG",value=lineaments_default,placeholder=lineaments_default)
    
    land_u_land_c_default = r"C:\images\6.jpg"
    land_u_land_c_path = st.text_input("Land Use Land COver IMG",value=land_u_land_c_default,placeholder=land_u_land_c_default)
    
    existing_bores_default = r"C:\images\-1.jpg"
    existing_bores_path = st.text_input("Existing_Bores_IMG",value=existing_bores_default,placeholder=existing_bores_default)
    
    grad_plan_default = r"C:\images\7.jpg"
    grad_plan_path = st.text_input("Gradient PLAN IMG",value=grad_plan_default,placeholder=grad_plan_default)
    
    
    res_grad_profiles_1_default =r"C:\images\8.jpg"
    res_grad_profiles_1_path = st.text_input("Resistivity Profile: 1",value=res_grad_profiles_1_default,placeholder=res_grad_profiles_1_default)
    res_grad_profiles_2_default =r"C:\images\9.jpg"
    res_grad_profiles_2_path = st.text_input("Resistivity Profile: 2",value=res_grad_profiles_2_default,placeholder=res_grad_profiles_2_default)
    res_contour_1_default =r"C:\images\10.jpg"
    res_contour_1_path = st.text_input("Resistivity_Contour Map",value=res_contour_1_default,placeholder=res_contour_1_default)
    res_contour_1_3d_default =r"C:\images\11.jpg"
    res_contour_1_3d_path = st.text_input("Resistivity_3D IMAGE",value=res_contour_1_3d_default,placeholder=res_contour_1_3d_default)
    admt_pqwt_plan_default =r"C:\images\12.jpg"
    
    
    admt_pqwt_plan_path = st.text_input("ADMT_PQWT PLAN IMAGE",value=admt_pqwt_plan_default,placeholder=admt_pqwt_plan_default)
    
    admt_l_1_default = r"C:\images\13.jpg"
    admt_l_1_path = st.text_input("ADMT_IMAGE-1",value=admt_l_1_default,placeholder=admt_l_1_default)
    
    admt_l_2_default = r"C:\images\14.jpg"
    admt_l_2_path = st.text_input("ADMT_IMAGE-2",value=admt_l_2_default,placeholder=admt_l_2_default)
    
    admt_l_3_default = r"C:\images\15.jpg"
    admt_l_3_path = st.text_input("ADMT_IMAGE-3",value=admt_l_3_default,placeholder=admt_l_3_default)
    
    admt_l_4_default = r"C:\images\16.jpg"
    admt_l_4_path = st.text_input("ADMT_IMAGE-4",value=admt_l_4_default,placeholder=admt_l_4_default)
    
    admt_l_5_default = r"C:\images\17.jpg"
    admt_l_5_path = st.text_input("ADMT_IMAGE-5",value=admt_l_5_default,placeholder=admt_l_5_default)
    
    admt_l_6_default = r"C:\images\35.jpg"
    admt_l_6_path = st.text_input("ADMT_IMAGE-6",value=admt_l_6_default,placeholder=admt_l_6_default)
    
    admt_l_7_default = r"C:\images\36.jpg"
    admt_l_7_path = st.text_input("ADMT_IMAGE-7",value=admt_l_7_default,placeholder=admt_l_7_default)
    
    admt_l_8_default = r"C:\images\37.jpg"
    admt_l_8_path = st.text_input("ADMT_IMAGE-8",value=admt_l_8_default,placeholder=admt_l_8_default)
    
    admt_l_9_default = r"C:\images\38.jpg"
    admt_l_9_path = st.text_input("ADMT_IMAGE-9",value=admt_l_9_default,placeholder=admt_l_9_default)
    
    admt_l_10_default = r"C:\images\39.jpg"
    admt_l_10_path = st.text_input("ADMT_IMAGE-10",value=admt_l_10_default,placeholder=admt_l_10_default)
    
    admt_l_11_default = r"C:\images\40.jpg"
    admt_l_11_path = st.text_input("ADMT_IMAGE-11",value=admt_l_11_default,placeholder=admt_l_11_default)
    
    admt_l_12_default = r"C:\images\41.jpg"
    admt_l_12_path = st.text_input("ADMT_IMAGE-12",value=admt_l_12_default,placeholder=admt_l_12_default)
    
    admt_l_13_default = r"C:\images\42.jpg"
    admt_l_13_path = st.text_input("ADMT_IMAGE-13",value=admt_l_13_default,placeholder=admt_l_13_default)
    
    admt_l_14_default = r"C:\images\43.jpg"
    admt_l_14_path = st.text_input("ADMT_IMAGE-14",value=admt_l_14_default,placeholder=admt_l_14_default)
    
    admt_l_15_default = r"C:\images\44.jpg"
    admt_l_15_path = st.text_input("ADMT_IMAGE-15",value=admt_l_15_default,placeholder=admt_l_15_default)
    
    admt_l_16_default = r"C:\images\45.jpg"
    admt_l_16_path = st.text_input("ADMT_IMAGE-16",value=admt_l_16_default,placeholder=admt_l_16_default)
    
    admt_l_17_default = r"C:\images\46.jpg"
    admt_l_17_path = st.text_input("ADMT_IMAGE-17",value=admt_l_17_default,placeholder=admt_l_17_default)
    
    admt_l_18_default = r"C:\images\47.jpg"
    admt_l_18_path = st.text_input("ADMT_IMAGE-18",value=admt_l_18_default,placeholder=admt_l_18_default)
   
    
    admt_l_19_default = r"C:\images\48.jpg"
    admt_l_19_path = st.text_input("ADMT_IMAGE-19",value=admt_l_19_default,placeholder=admt_l_19_default)
    
    admt_l_20_default = r"C:\images\49.jpg"
    admt_l_20_path = st.text_input("ADMT_IMAGE-20",value=admt_l_20_default,placeholder=admt_l_20_default)
    
    recom_points_default = r"C:\images\18.jpg"
    recom_points_path = st.text_input("Recommended Points IMG",value=recom_points_default,placeholder=recom_points_default)
    
    field_pic_default = r"C:\images\19.jpg"
    field_pic_path = st.text_input("Field Picture",value=field_pic_default,placeholder=field_pic_default)
    
    pqwt_l_1_default =r"C:\images\20.jpg"
    pqwt_l_1_path = st.text_input("PQWT IMG-1",value= pqwt_l_1_default,placeholder=pqwt_l_1_default)
    
    pqwt_l_2_default =r"C:\images\21.jpg"
    pqwt_l_2_path = st.text_input("PQWT IMG-2",value= pqwt_l_2_default,placeholder=pqwt_l_2_default)
    
    pqwt_l_3_default =r"C:\images\22.jpg"
    pqwt_l_3_path = st.text_input("PQWT IMG-3",value= pqwt_l_3_default,placeholder=pqwt_l_3_default)
    
    pqwt_l_4_default =r"C:\images\23.jpg"
    pqwt_l_4_path = st.text_input("PQWT IMG-4",value= pqwt_l_4_default,placeholder=pqwt_l_4_default)
    
    pqwt_l_5_default =r"C:\images\24.jpg"
    pqwt_l_5_path = st.text_input("PQWT IMG-5",value= pqwt_l_5_default,placeholder=pqwt_l_5_default)
    
    pqwt_l_6_default =r"C:\images\25.jpg"
    pqwt_l_6_path = st.text_input("PQWT IMG-6",value= pqwt_l_6_default,placeholder=pqwt_l_6_default)
    
    pqwt_l_7_default =r"C:\images\26.jpg"
    pqwt_l_7_path = st.text_input("PQWT IMG-7",value= pqwt_l_7_default,placeholder=pqwt_l_7_default)
    
    pqwt_l_8_default =r"C:\images\27.jpg"
    pqwt_l_8_path = st.text_input("PQWT IMG-8",value= pqwt_l_8_default,placeholder=pqwt_l_8_default)
    
    pqwt_l_9_default =r"C:\images\28.jpg"
    pqwt_l_9_path = st.text_input("PQWT IMG-9",value= pqwt_l_9_default,placeholder=pqwt_l_9_default)
    
    pqwt_l_10_default =r"C:\images\29.jpg"
    pqwt_l_10_path = st.text_input("PQWT IMG-10",value= pqwt_l_10_default,placeholder=pqwt_l_10_default)
    
    pqwt_l_11_default =r"C:\images\30.jpg"
    pqwt_l_11_path = st.text_input("PQWT IMG-11",value= pqwt_l_11_default,placeholder=pqwt_l_11_default)
    
    pqwt_l_12_default =r"C:\images\31.jpg"
    pqwt_l_12_path = st.text_input("PQWT IMG-12",value= pqwt_l_12_default,placeholder=pqwt_l_12_default)
    
    pqwt_l_13_default =r"C:\images\32.jpg"
    pqwt_l_13_path = st.text_input("PQWT IMG-13",value= pqwt_l_13_default,placeholder=pqwt_l_13_default)
    
    pqwt_l_14_default =r"C:\images\33.jpg"
    pqwt_l_14_path = st.text_input("PQWT IMG-14",value= pqwt_l_14_default,placeholder=pqwt_l_14_default)
    
    pqwt_l_15_default =r"C:\images\34.jpg"
    pqwt_l_15_path = st.text_input("PQWT IMG-15",value= pqwt_l_15_default,placeholder=pqwt_l_15_default)
    
    
    if st.button("Submit"):
        # Validation
        if not terrain_type:
            st.error("Terrain type is required.")
            return
        
        if not geomorphic_unit:
            st.error("Geomorphology is required")
            return
        if not type_of_land:
            st.error("Type of Land is required")
            return
        
        if not soil_type:
            st.error("Type of soil is required")
            return
            
        if not des_soil:
            st.error("SOil description is required")
            return
            
        if not year_rainfall:
            st.error("year_rainfall is required")
            return
            
        if not water_quality:
            st.error("Water quality is required")
            return
       
        if not grad_trav_direction:
            st.error("Gradient line direction is required")
            return
            
            
        if not grad_station_int:
            st.error("Grad interval is required")
            return
        
        if not c1c2:
            st.error("C1C2 spacing is required")
        
        if not admt_electrode_int:
            st.error("ADMT_Electode spcaing is required")
            
            
        #if highest_elevation and highest_elevation_num is None:
        #    st.error("Highest elevation must be numeric.")
        #    return
            
        if min_lat and min_lat_num is None:
            st.error("min_lat must be numeric.")
            return
            
        if max_lat and max_lat_num is None:
            st.error("max_lat must be numeric.")
            return
            
        if min_lon and min_lon_num is None:
            st.error("min_lon must be numeric.")
            return
        
        
        if max_lon and max_lon_num is None:
            st.error("max_lon must be numeric.")
            return
            
        if figure_number is None:
            st.error("Wrong entry")
            return
            
        if overlying_materials is None:
            st.error ("pls type overlying material")
            return
            
 

        geology_img_path = geology_img_path.strip()
        home_page_path = home_page_path.strip()
        study_area_path = study_area_path.strip()
        gw_prospects_path = gw_prospects_path.strip()
        drainage_pattern_path = drainage_pattern_path.strip()
        lineaments_path = lineaments_path.strip()
        land_u_land_c_path = land_u_land_c_path.strip()
        existing_bores_path = existing_bores_path.strip()
        grad_plan_path = grad_plan_path.strip()
        res_grad_profiles_1_path = res_grad_profiles_1_path.strip()
        res_grad_profiles_2_path = res_grad_profiles_2_path.strip()
        res_contour_1_path = res_contour_1_path.strip()
        res_contour_1_3d_path = res_contour_1_3d_path.strip()
        admt_pqwt_plan_path = admt_pqwt_plan_path.strip()
        admt_l_1_path = admt_l_1_path.strip()
        admt_l_2_path = admt_l_2_path.strip()
        admt_l_3_path = admt_l_3_path.strip()
        admt_l_4_path = admt_l_4_path.strip()
        admt_l_5_path = admt_l_5_path.strip()
        admt_l_6_path = admt_l_6_path.strip()
        admt_l_7_path = admt_l_7_path.strip()
        admt_l_8_path = admt_l_8_path.strip()
        admt_l_9_path = admt_l_9_path.strip()
        admt_l_10_path = admt_l_10_path.strip()
        admt_l_11_path = admt_l_11_path.strip()
        admt_l_12_path = admt_l_12_path.strip()
        admt_l_13_path = admt_l_13_path.strip()
        admt_l_14_path = admt_l_14_path.strip()
        admt_l_15_path = admt_l_15_path.strip()
        admt_l_16_path = admt_l_16_path.strip()
        admt_l_17_path = admt_l_17_path.strip()
        admt_l_18_path = admt_l_18_path.strip()
        admt_l_19_path = admt_l_19_path.strip()
        admt_l_20_path = admt_l_20_path.strip()
        recom_points_path = recom_points_path.strip()
        field_pic_path = field_pic_path.strip()
        pqwt_l_1_path = pqwt_l_1_path.strip()
        pqwt_l_2_path = pqwt_l_2_path.strip()
        pqwt_l_3_path = pqwt_l_3_path.strip()
        pqwt_l_4_path = pqwt_l_4_path.strip()
        pqwt_l_5_path = pqwt_l_5_path.strip()
        pqwt_l_6_path = pqwt_l_6_path.strip()
        pqwt_l_7_path = pqwt_l_7_path.strip()
        pqwt_l_8_path = pqwt_l_8_path.strip()
        pqwt_l_9_path = pqwt_l_9_path.strip()
        pqwt_l_10_path = pqwt_l_10_path.strip()
        pqwt_l_11_path = pqwt_l_11_path.strip()
        pqwt_l_12_path = pqwt_l_12_path.strip()
        pqwt_l_13_path = pqwt_l_13_path.strip()
        pqwt_l_14_path = pqwt_l_14_path.strip()
        pqwt_l_15_path = pqwt_l_15_path.strip()
        
        data = {
            "terrain_type": terrain_type,
            "geomorphic_unit":geomorphic_unit,
            "date": date_selected,
        #    "highest_elevation": highest_elevation_num if highest_elevation else "",
            "geology_img": geology_img_path,
            "figure_number": figure_number,
            "village":village,
            "mandal": mandal,
            "district": district,
            "state":state,
            "pin":pin,
            "area":area,
            "client_name":client_name,
            "min_lat":min_lat,
            "max_lat":max_lat,
            "min_lon":min_lon,
            "max_lon":max_lon,
            "weathering_depth":weathering_depth,
            "comparison_depth":comparison_depth,
            "type_of_land":type_of_land,
            "soil_type":soil_type,
            "des_soil":des_soil,
            "year_rainfall":year_rainfall,
            "water_quality":water_quality,
            "grad_trav_direction":grad_trav_direction,
            "grad_station_int":grad_station_int,
            "c1c2":c1c2,
            "admt_electrode_int":admt_electrode_int,
            "home_page": home_page_path,
            "study_area":study_area_path,
            "gw_prospects":gw_prospects_path,
            "drainage_pattern":drainage_pattern_path,
            "lineaments":lineaments_path,
            "land_u_land_c": land_u_land_c_path,
            "existing_bores": existing_bores_path,
            "grad_plan":grad_plan_path,
            "res_grad_profiles_1": res_grad_profiles_1_path,
            "res_grad_profiles_2": res_grad_profiles_2_path,
            "res_contour_1":res_contour_1_path,
            "res_contour_1_3d":res_contour_1_3d_path,
            "admt_pqwt_plan":admt_pqwt_plan_path,
            "admt_l_1": admt_l_1_path,
            "admt_l_2": admt_l_2_path,
            "admt_l_3": admt_l_3_path,
            "admt_l_4": admt_l_4_path,
            "admt_l_5": admt_l_5_path,
            "admt_l_6": admt_l_6_path,
            "admt_l_7": admt_l_7_path,
            "admt_l_8": admt_l_8_path,
            "admt_l_9": admt_l_9_path,
            "admt_l_10": admt_l_10_path,
            "admt_l_11": admt_l_11_path,
            "admt_l_12": admt_l_12_path,
            "admt_l_13": admt_l_13_path,
            "admt_l_14": admt_l_14_path,
            "admt_l_15": admt_l_15_path,
            "admt_l_16": admt_l_16_path,
            "admt_l_17": admt_l_17_path,
            "admt_l_18": admt_l_18_path,
            "admt_l_19": admt_l_19_path,
            "admt_l_20": admt_l_20_path,
            "recom_points":recom_points_path,
            "field_pic": field_pic_path,
            "to_nearest_town":to_nearest_town,
            "nearest_town":nearest_town,
            "to_nearest_city":to_nearest_city,
            "nearest_city":nearest_city,
            "rain_fall":rain_fall,
            "temp":temp,
            "yield_from_prospects":yield_from_prospects,
            "no_grad_trav": no_grad_trav,
            "grad_trav_name":grad_trav_name,
            "res_highs":res_highs,
            "res_high_extends":res_high_extends,
            "res_low":res_low,
            "res_low_extends":res_low_extends,
            "res_relief":res_relief,
            "recom_bores":recom_bores,
            "final_recom_points":final_recom_points,
            "recom_points_order":recom_points_order,
            "water_zone_depths":water_zone_depths,
            "yield":yield_,
            "considerable_depths":considerable_depths,
            "ref_no":ref_no,
            "res_low_trend_des": res_low_trend_des,
            "highest_elevation":highest_elevation,
            "lowest_elevation":lowest_elevation,
            "pqwt_l_1": pqwt_l_1_path,
            "pqwt_l_2": pqwt_l_2_path,
            "pqwt_l_3": pqwt_l_3_path,
            "pqwt_l_4": pqwt_l_4_path,
            "pqwt_l_5": pqwt_l_5_path,
            "pqwt_l_6": pqwt_l_6_path,
            "pqwt_l_7": pqwt_l_7_path,
            "pqwt_l_8": pqwt_l_8_path,
            "pqwt_l_9": pqwt_l_9_path,
            "pqwt_l_10": pqwt_l_10_path,
            "pqwt_l_11": pqwt_l_11_path,
            "pqwt_l_12": pqwt_l_12_path,
            "pqwt_l_13": pqwt_l_13_path,
            "pqwt_l_14": pqwt_l_14_path,
            "pqwt_l_15": pqwt_l_15_path,
            "hydrogeology": h_geo
            
        }

        if "data_df" not in st.session_state:
            st.session_state.data_df = pd.DataFrame(columns=list(data.keys()))

        new_row = pd.DataFrame([data])
        st.session_state.data_df = pd.concat([st.session_state.data_df, new_row], ignore_index=True)

        st.success("Data saved. Image path captured.")

        st.write("Current Data:")
        st.dataframe(st.session_state.data_df)

        excel_bytes = save_to_excel(st.session_state.data_df, file_path="data.xlsx")
        
        st.download_button(
            label="Download Excel",
            data=excel_bytes,
            file_name="data.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

if __name__ == "__main__":
    main()
